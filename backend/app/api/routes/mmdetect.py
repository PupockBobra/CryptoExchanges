"""
REST API for the MM-presence estimator (SPB perps).

Everything here reads stored order-book snapshots — no Finam call happens on any
of these paths, so re-running the detector with different thresholds costs a
query, not market access.

Endpoints
---------
GET /api/mmdetect/instruments  → capture coverage per instrument
GET /api/mmdetect/analyze      → full estimate + evidence + charts for one
GET /api/mmdetect/summary      → the comparison table across all instruments
GET /api/mmdetect/export.csv   → summary + confirmed clusters as CSV
GET /api/mmdetect/export.xlsx  → the summary table as the page shows it
"""

import asyncio
import csv
import io
import logging

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.api.cache import ttl_cache
from app.mmdetect.config import (
    CORRIDORS,
    DEFAULTS,
    MMD_TICKERS,
    SESSION_MODE_LABELS,
    DetectParams,
)
from app.db.timescale import get_latest_usdrub
from app.mmdetect.service import analyze_symbol, coverage, stride_for, window
from app.spb.config import SPB_GROUP_ORDER, SPB_GROUPS, SPB_LOTS, SPB_NAMES

log = logging.getLogger(__name__)
router = APIRouter()


def _params(persistence_min: float, volume_tol: float, symmetry_tol: float,
            bin_steps: int | None, min_cluster_volume: float = DEFAULTS.min_cluster_volume,
            search_radius_pct: float = DEFAULTS.search_radius_pct) -> DetectParams:
    """Thresholds from the query string, falling back to the configured
    defaults.  ``bin_steps=0`` means "derive it from the data" (the page's
    slider sends 0 for its Auto position)."""
    return DetectParams(
        persistence_min=persistence_min,
        volume_tol=volume_tol,
        symmetry_tol=symmetry_tol,
        bin_steps=bin_steps or None,
        min_cluster_volume=min_cluster_volume,
        search_radius_pct=search_radius_pct,
        corridors=CORRIDORS,
    )


def _money(pair: dict, ticker: str, usdrub: float | None) -> dict:
    """A quoter's resting size in money.

    ``notional_usd`` from the core is per contract-unit, so the venue's contract
    multiplier is applied here — this is the layer that knows the instrument.
    Roubles come from the same ``moex_fx_rates`` USDRUBF rate the rest of the app
    converts with, so every page states money in one currency basis.
    """
    n = pair.get("notional_usd")
    if n is None:
        return {"volume_usd": None, "volume_rub": None}
    usd = n * SPB_LOTS.get(ticker, 1.0)
    return {"volume_usd": usd, "volume_rub": (usd * usdrub) if usdrub else None}


def _check_ticker(ticker: str) -> None:
    if ticker not in MMD_TICKERS:
        raise HTTPException(404, f"unknown instrument {ticker!r}")


@router.get("/instruments")
@ttl_cache()
async def get_instruments(hours: float = Query(6, gt=0, le=48)):
    """Instruments being captured, with how complete the record is over the
    window.  ``miss_ratio`` is the share of trading-hours grid points with no
    stored snapshot — the qualifier on every persistence number downstream."""
    ts_from, ts_to = window(hours)
    cov = await coverage(ts_from, ts_to)
    return {
        "from": ts_from.isoformat(),
        "to": ts_to.isoformat(),
        "modes": SESSION_MODE_LABELS,
        "defaults": {
            "persistence_min": DEFAULTS.persistence_min,
            "volume_tol": DEFAULTS.volume_tol,
            "symmetry_tol": DEFAULTS.symmetry_tol,
            "min_cluster_volume": DEFAULTS.min_cluster_volume,
            "search_radius_pct": DEFAULTS.search_radius_pct,
            "corridors": list(CORRIDORS),
        },
        "instruments": [cov[t] for t in MMD_TICKERS],
    }


@router.get("/analyze")
@ttl_cache()
async def get_analyze(
    ticker: str,
    hours: float = Query(6, gt=0, le=48),
    mode: str = "all",
    persistence_min: float = Query(DEFAULTS.persistence_min, ge=0, le=1),
    volume_tol: float = Query(DEFAULTS.volume_tol, ge=0, lt=1),
    symmetry_tol: float = Query(DEFAULTS.symmetry_tol, ge=0, le=1),
    bin_steps: int = Query(0, ge=0, le=1000),
    min_cluster_volume: float = Query(DEFAULTS.min_cluster_volume, ge=0),
    search_radius_pct: float = Query(DEFAULTS.search_radius_pct, gt=0, le=0.2),
    heatmap_cols: int = Query(240, ge=20, le=600),
):
    """Everything the instrument's card shows: the estimate, the clusters behind
    it, the heat map, the spread series and the persistence profile."""
    _check_ticker(ticker)
    ts_from, ts_to = window(hours)
    usdrub = await get_latest_usdrub()
    out = await analyze_symbol(
        ticker, ts_from, ts_to,
        _params(persistence_min, volume_tol, symmetry_tol, bin_steps, min_cluster_volume,
                    search_radius_pct),
        mode=mode, heatmap_cols=heatmap_cols,
    )
    for p in out["result"]["pairs"]:
        p.update(_money(p, ticker, usdrub))
    out["usdrub"] = usdrub
    return out


@router.get("/summary")
@ttl_cache()
async def get_summary(
    hours: float = Query(6, gt=0, le=48),
    mode: str = "all",
    persistence_min: float = Query(DEFAULTS.persistence_min, ge=0, le=1),
    volume_tol: float = Query(DEFAULTS.volume_tol, ge=0, lt=1),
    symmetry_tol: float = Query(DEFAULTS.symmetry_tol, ge=0, le=1),
    bin_steps: int = Query(0, ge=0, le=1000),
    min_cluster_volume: float = Query(DEFAULTS.min_cluster_volume, ge=0),
    search_radius_pct: float = Query(DEFAULTS.search_radius_pct, gt=0, le=0.2),
):
    """One row per instrument, same thresholds for all of them.

    Sampled at a stride (see ``stride_for``) so the whole table is one
    interactive request; the per-instrument view uses the full 5-second grid, so
    the two can differ slightly — the stride is reported in every row.
    """
    params = _params(persistence_min, volume_tol, symmetry_tol, bin_steps, min_cluster_volume,
                    search_radius_pct)
    analysed, ctx = await _summary_rows(hours, mode, params)
    ts_from, ts_to, stride, usdrub = ctx["from"], ctx["to"], ctx["stride"], ctx["usdrub"]

    rows = []
    for ticker, res, cv in analysed:
        corr = res.get("corridors") or {}
        rows.append({
            "ticker": ticker,
            "name": SPB_NAMES.get(ticker, ticker),
            "group": SPB_GROUPS.get(ticker, ""),
            "n_snapshots": res["n_snapshots"],
            "enough_data": res["enough_data"],
            "spread_observed_bps": res["spread_observed"]["median"],
            "spread_mm_bps": res["spread_mm"]["median"],
            # Total resting across ALL confirmed quoters, plus the largest single
            # one: a book can hold several makers (AMD does), and one averaged
            # figure would describe none of them.
            "mm_volume": (res["mm_volume"] or {}).get("total"),
            "mm_volume_largest": (res["mm_volume"] or {}).get("largest"),
            "n_pairs": (res["mm_volume"] or {}).get("n_pairs", 0),
            # Where each side of the pair rests.  The two are reported separately
            # because they legitimately differ — the obligation ties the sizes,
            # not the distances — and a pair resting deep explains an MM spread
            # far wider than the observed one.
            "dist_bid_steps": (res["pairs"][0]["dist_bid_steps"] if res["pairs"] else None),
            "dist_ask_steps": (res["pairs"][0]["dist_ask_steps"] if res["pairs"] else None),
            "valid_match_share": res["valid_match_share"],
            # Every detected quoter, not just the aggregate: a book with three
            # makers of 140, 50 and 10 is a different fact from one maker of 200,
            # and a total cannot tell the two apart.  Trimmed to the fields the
            # table shows; the full per-quoter distribution stays in /analyze.
            "quoters": [{
                "volume_bid": p["volume_bid"],
                "volume_ask": p["volume_ask"],
                "volume_two_sided": p["volume_two_sided"],
                "dist_bid_steps": p["dist_bid_steps"],
                "dist_ask_steps": p["dist_ask_steps"],
                "spread_bps": p["spread_bps"]["median"],
                "presence_bid": p["presence_bid"],
                "presence_ask": p["presence_ask"],
                "alone_bid": p["alone_bid"],
                "alone_ask": p["alone_ask"],
                "match_share": p["match_share"],
                **_money(p, ticker, usdrub),
            } for p in res["pairs"]],
            "corridors": {
                k: {"two_sided": v["two_sided"]["median"],
                    "two_sided_usd": v["two_sided_usd"]["median"],
                    "truncated_share": v["truncated_share"]}
                for k, v in corr.items()
            },
            "miss_ratio": cv["miss_ratio"],
            "stride_sec": stride,
        })
    return {"from": ts_from.isoformat(), "to": ts_to.isoformat(), "mode": mode,
            "corridors": list(CORRIDORS), "usdrub": usdrub, "rows": rows}


@router.get("/export.csv")
async def get_export(
    hours: float = Query(6, gt=0, le=48),
    mode: str = "all",
    persistence_min: float = Query(DEFAULTS.persistence_min, ge=0, le=1),
    volume_tol: float = Query(DEFAULTS.volume_tol, ge=0, lt=1),
    symmetry_tol: float = Query(DEFAULTS.symmetry_tol, ge=0, le=1),
    bin_steps: int = Query(0, ge=0, le=1000),
    min_cluster_volume: float = Query(DEFAULTS.min_cluster_volume, ge=0),
    search_radius_pct: float = Query(DEFAULTS.search_radius_pct, gt=0, le=0.2),
):
    """Summary table and the confirmed clusters behind it, as one CSV.

    Two sections in one file on purpose: a summary row is only as good as the
    clusters it came from, and shipping them apart invites quoting the estimate
    without its evidence.
    """
    ts_from, ts_to = window(hours)
    params = _params(persistence_min, volume_tol, symmetry_tol, bin_steps, min_cluster_volume,
                    search_radius_pct)
    stride = stride_for(hours)
    cov = await coverage(ts_from, ts_to)

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"MM presence estimate (SPB), {ts_from.isoformat()} .. {ts_to.isoformat()}, "
                f"mode={mode}, persistence>={persistence_min}, vol_tol={volume_tol}, "
                f"sym_tol={symmetry_tol}, stride={stride}s"])
    w.writerow([])
    w.writerow(["Instrument", "Name", "Snapshots", "Miss ratio",
                "Observed spread, bps", "MM spread, bps", "MM volume, contracts",
                "Quoters found", "Valid match share",
                "Bid distance, steps", "Ask distance, steps",
                *[f"Two-sided depth ±{float(c) * 100:g}%" for c in CORRIDORS],
                *[f"Truncated share ±{float(c) * 100:g}%" for c in CORRIDORS]])

    cluster_rows = []
    for ticker in MMD_TICKERS:
        res = (await analyze_symbol(ticker, ts_from, ts_to, params, mode=mode,
                                    with_heatmap=False, stride_sec=stride))["result"]
        corr = res.get("corridors") or {}
        w.writerow([
            ticker, SPB_NAMES.get(ticker, ticker), res["n_snapshots"],
            _r(cov[ticker]["miss_ratio"], 4),
            _r(res["spread_observed"]["median"], 2),
            _r(res["spread_mm"]["median"], 2),
            _r((res["mm_volume"] or {}).get("total"), 2),
            (res["mm_volume"] or {}).get("n_pairs", 0),
            _r(res["valid_match_share"], 4),
            _r(res["pairs"][0]["dist_bid_steps"] if res["pairs"] else None, 1),
            _r(res["pairs"][0]["dist_ask_steps"] if res["pairs"] else None, 1),
            *[_r((corr.get(str(c)) or {}).get("two_sided", {}).get("median"), 2) for c in CORRIDORS],
            *[_r((corr.get(str(c)) or {}).get("truncated_share"), 4) for c in CORRIDORS],
        ])
        for p in res["pairs"]:
            cluster_rows.append([
                ticker,
                _r(p["volume_bid"], 2), _r(p["volume_ask"], 2),
                _r(p["volume_two_sided"], 2),
                _r(p["dist_bid_steps"], 1), _r(p["dist_ask_steps"], 1),
                _r(p["presence_bid"], 4), _r(p["presence_ask"], 4),
                _r(p["spread_bps"]["median"], 2), _r(p["match_share"], 4),
            ])

    w.writerow([])
    w.writerow(["Confirmed two-sided clusters"])
    w.writerow(["Instrument", "Volume bid", "Volume ask", "Two-sided volume",
                "Bid distance, steps", "Ask distance, steps",
                "Presence bid", "Presence ask", "Spread, bps", "Both sides share"])
    w.writerows(cluster_rows)

    buf.seek(0)
    fname = f"mm_presence_{ts_to.date()}_{mode}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _r(v, nd):
    return "" if v is None else round(float(v), nd)


async def _summary_rows(hours: float, mode: str, params: DetectParams) -> tuple[list[dict], dict]:
    """The summary table's rows plus the header context, shared by the page, the
    CSV and the Excel export so all three can never drift apart."""
    ts_from, ts_to = window(hours)
    stride = stride_for(hours)
    cov = await coverage(ts_from, ts_to)
    usdrub = await get_latest_usdrub()
    # Concurrently: each instrument is an independent DB read followed by CPU
    # work, and run one-by-one the reads alone took ~8 s of the request.  The
    # analysis itself still serialises on the GIL — the point is to overlap the
    # waiting, not to parallelise the maths.  The cap keeps 20 simultaneous
    # readers off the connection pool.
    sem = asyncio.Semaphore(5)

    async def one(ticker: str):
        async with sem:
            res = (await analyze_symbol(ticker, ts_from, ts_to, params, mode=mode,
                                        with_heatmap=False, stride_sec=stride))["result"]
        return ticker, res, cov[ticker]

    rows = list(await asyncio.gather(*(one(t) for t in MMD_TICKERS)))
    return rows, {"from": ts_from, "to": ts_to, "stride": stride, "usdrub": usdrub}


@router.get("/export.xlsx")
async def get_export_xlsx(
    hours: float = Query(6, gt=0, le=48),
    mode: str = "all",
    persistence_min: float = Query(DEFAULTS.persistence_min, ge=0, le=1),
    volume_tol: float = Query(DEFAULTS.volume_tol, ge=0, lt=1),
    symmetry_tol: float = Query(DEFAULTS.symmetry_tol, ge=0, le=1),
    min_cluster_volume: float = Query(DEFAULTS.min_cluster_volume, ge=0),
    search_radius_pct: float = Query(DEFAULTS.search_radius_pct, gt=0, le=0.2),
):
    """The summary table as a real .xlsx — one row per detected quoter.

    Carries the thresholds it was produced with in the first rows: the numbers
    below them are only meaningful together with the settings that produced them,
    and a spreadsheet outlives the page state it was exported from.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    params = _params(persistence_min, volume_tol, symmetry_tol, None, min_cluster_volume,
                     search_radius_pct)
    rows, ctx = await _summary_rows(hours, mode, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "MM presence"

    ws.append(["Присутствие маркет-мейкеров — СПБ Биржа"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Окно: {ctx['from']:%d.%m.%Y %H:%M} – {ctx['to']:%d.%m.%Y %H:%M} UTC, "
               f"режим: {SESSION_MODE_LABELS.get(mode, mode)}, шаг сетки: {ctx['stride']} с"])
    ws.append([f"Пороги: персистентность ≥ {persistence_min:.0%}, допуск по объёму ±{volume_tol:.0%}, "
               f"симметрия {symmetry_tol:.0%}, мин. объём {min_cluster_volume:g}, "
               f"радиус {search_radius_pct:.2%}"
               + (f", курс USD/RUB {ctx['usdrub']:.2f}" if ctx["usdrub"] else ", курс USD/RUB недоступен")])
    ws.append(["Оценка присутствия ММ, не измерение: алго-участники котируют так же; "
               "несколько ММ одного размера сольются в одного; айсберги видны частично."])
    ws.append([])

    head = ["Инструмент", "Снимков", "Пропуск", "Спред набл., б.п.", "Объём ММ всего, контр."]
    head += [f"±{c * 100:g}%, контр." for c in CORRIDORS]
    head += ["Котировщик", "Объём, контр.", "Объём, $", "Объём, ₽",
             "Удаление bid, ш.", "Удаление ask, ш.", "Спред ММ, б.п.",
             "Стоял bid", "Стоял ask", "Один на уровне bid", "Один на уровне ask",
             "Обе стороны"]
    ws.append(head)
    hrow = ws.max_row
    n_base = 5 + len(CORRIDORS)      # columns describing the instrument itself
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ticker, res, cv in rows:
        corr = res.get("corridors") or {}
        base = [
            SPB_NAMES.get(ticker, ticker), res["n_snapshots"],
            cv["miss_ratio"], res["spread_observed"]["median"],
            (res["mm_volume"] or {}).get("total"),
            *[(corr.get(str(c)) or {}).get("two_sided", {}).get("median") for c in CORRIDORS],
        ]
        pairs = res["pairs"] or [None]
        for i, p in enumerate(pairs):
            lead = base if i == 0 else [None] * n_base
            if p is None:
                ws.append([*lead, "не подтверждён"])
                continue
            money = _money(p, ticker, ctx["usdrub"])
            ws.append([*lead, f"№{i + 1}",
                       p["volume_two_sided"], money["volume_usd"], money["volume_rub"],
                       p["dist_bid_steps"], p["dist_ask_steps"], p["spread_bps"]["median"],
                       p["presence_bid"], p["presence_ask"],
                       p["alone_bid"], p["alone_ask"], p["match_share"]])
        if len(pairs) > 1:      # merge the instrument's own figures down its group
            top, bottom = ws.max_row - len(pairs) + 1, ws.max_row
            for c in range(1, n_base + 1):
                ws.merge_cells(start_row=top, start_column=c, end_row=bottom, end_column=c)
                ws.cell(row=top, column=c).alignment = Alignment(vertical="center")

    pct_cols = {3} | {n_base + 8 + k for k in range(5)}      # shares → percent format
    for row in ws.iter_rows(min_row=hrow + 1):
        for cell in row:
            if not isinstance(cell.value, (int, float)):
                continue
            if cell.column in pct_cols:
                cell.number_format = "0%"
            elif cell.column in (n_base + 3, n_base + 4):
                cell.number_format = "#,##0"
            else:
                cell.number_format = "#,##0.00"

    widths = [26, 9, 9, 11, 12] + [11] * len(CORRIDORS) + [12, 12, 13, 14, 12, 12, 11, 10, 10, 12, 12, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"mm_presence_{ctx['to'].date()}_{mode}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
